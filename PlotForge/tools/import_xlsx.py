# -*- coding: utf-8 -*-
"""
Plot plan 요구사항 xlsx → Sitelayout JSON '초안' 변환기 (오프라인 개발 도구).

역할: 기계 추출이 확실한 것만 자동화한다 —
  Legend 시트  → buildingTypes 스텁 (이름·번호. 치수는 기본값, 사람이 수정)
  Layout 시트  → _blockDrafts (블록명·멤버 번호·요구문장 원문)
규칙 수치·hard/soft 판단은 자동 번역하지 않는다(오번역이 조용한 hard 오류가 되므로).
출력 JSON의 '_'로 시작하는 필드는 앱 restore()가 무시하는 참고용 초안 데이터다.

사용법:
  python tools/import_xlsx.py "Plot plan requirement.xlsx" -o demos/draft.json
  python tools/import_xlsx.py <파일> --legend Legend --layout Layout
"""
import argparse
import json
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl이 필요합니다: python -m pip install openpyxl")

PALETTE = ['#e57373', '#64b5f6', '#81c784', '#ffb74d', '#ba68c8',
           '#4db6ac', '#f06292', '#a1887f', '#90a4ae', '#dce775']


def read_legend(ws):
    """번호(B열) → 시설물명(C열). 이름 없는 번호는 건너뜀."""
    items = {}
    for row in ws.iter_rows(min_row=1):
        no, name = None, None
        for cell in row:
            if cell.column_letter == 'B' and isinstance(cell.value, (int, float)):
                no = int(cell.value)
            if cell.column_letter == 'C' and isinstance(cell.value, str) and cell.value.strip():
                name = cell.value.strip()
        if no is not None and name:
            items[no] = name
    return items


def read_layout(ws):
    """병합 셀 구조의 Layout 시트 → [{block, itemNos, requirement}]."""
    # 병합 해제 값 채우기: (row, col) → top-left 값
    merged_val = {}
    for rng in ws.merged_cells.ranges:
        tl = ws.cell(rng.min_row, rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merged_val[(r, c)] = tl

    def val(r, c):
        v = ws.cell(r, c).value
        if v is None:
            v = merged_val.get((r, c))
        return v

    COL = {'block': 2, 'no': 3, 'desc': 4, 'req': 5}  # B~E
    blocks = {}
    order = []
    for r in range(3, ws.max_row + 1):  # 헤더(2행) 다음부터
        block = val(r, COL['block'])
        no = val(r, COL['no'])
        req = val(r, COL['req'])
        if not isinstance(block, str) or not block.strip():
            continue
        block = block.strip()
        if block not in blocks:
            blocks[block] = {'block': block, 'itemNos': [], 'requirement': ''}
            order.append(block)
        if isinstance(no, (int, float)):
            blocks[block]['itemNos'].append(int(no))
        if isinstance(req, str) and req.strip() and not blocks[block]['requirement']:
            blocks[block]['requirement'] = req.strip()
    return [blocks[b] for b in order]


def build_draft(legend, layouts):
    """serialize 호환 초안 JSON. 치수·규칙은 사람이 완성."""
    skip_markers = ('고려하지 않음', '배치하지 않음')
    types = []
    for i, (no, name) in enumerate(sorted(legend.items())):
        types.append({
            'id': f't_{no}',
            'name': name,
            'w': 2, 'h': 2,  # TODO: 실제 치수(셀)로 수정 — cellSize(m/셀) 기준
            'color': PALETTE[i % len(PALETTE)],
            '_legendNo': no,
        })
    block_drafts = []
    for lay in layouts:
        skipped = any(m in lay['requirement'] for m in skip_markers)
        block_drafts.append({
            'name': lay['block'],
            'memberTypeIds': [f't_{n}' for n in lay['itemNos']],
            'requirementText': lay['requirement'],
            'skipped': skipped,  # xlsx가 미배치로 명시한 블록
            '_todoRules': _hint_rules(lay['requirement']),
        })
    return {
        'schemaVersion': 2,
        'site': None, 'cellSize': 2, 'maxCoverage': None,
        'roadNetwork': [], 'roadPattern': 'loop', 'entryCell': None,
        'buildingTypes': types,
        'rules': [],       # 규칙은 사람이 앱/에디터에서 완성 (자동 번역하지 않음)
        'sequence': [], 'alternatives': [], 'activeAltId': None,
        '_blockDrafts': block_drafts,
        '_note': '초안: 치수(w/h)와 rules는 요구문장(requirementText)을 보고 사람이 완성할 것. '
                 '_todoRules는 키워드 힌트일 뿐 검증된 번역이 아님.',
    }


def _hint_rules(text):
    """요구문장 → 규칙 '힌트' 목록 (참고용 키워드 매칭. 번역 아님)."""
    hints = []
    if not text:
        return hints
    for pat, hint in [
        (r'붙여서|가까이|가깝게', 'distanceTo(max) 후보'),
        (r'(\d+(?:\.\d+)?)\s*m\s*(?:이상\s*)?이격', '거리범위 min 후보 (m값 추출)'),
        (r'(\d+(?:\.\d+)?)\s*m\s*이내', '거리범위 max 후보 (m값 추출)'),
        (r'바람\s*방향', 'windSide(down) 후보'),
        (r'바람.*등지', 'windSide(up) 후보'),
        (r'중앙', 'centerOf 후보'),
        (r'중간에\s*배치', 'between 후보'),
        (r'[Ff]ence|경계|외곽', 'nearRoad(basis:fence) 후보'),
        (r'도로\s*배치|두르는', '둘레도로(ringRoad) 후보'),
        (r'[Uu]ser가\s*위치\s*지정', '수동 배치/앵커 후보'),
        (r'폭\s*(\d+(?:\.\d+)?)\s*m', 'corridor 폭 (m값 추출)'),
    ]:
        m = re.search(pat, text)
        if m:
            hints.append(hint + (f' [{m.group(1)}m]' if m.groups() and m.group(1) else ''))
    return hints


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx')
    ap.add_argument('-o', '--out', default='demos/draft.json')
    ap.add_argument('--legend', default='Legend')
    ap.add_argument('--layout', default='Layout')
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    legend = read_legend(wb[args.legend])
    layouts = read_layout(wb[args.layout])
    draft = build_draft(legend, layouts)

    import os
    os.makedirs(os.path.dirname(args.out) or '.', exist_ok=True)
    with open(args.out, 'w', encoding='utf-8') as f:
        json.dump(draft, f, ensure_ascii=False, indent=2)
    print(f'OK: 타입 {len(draft["buildingTypes"])}개, 블록 초안 {len(draft["_blockDrafts"])}개 '
          f'(미배치 표시 {sum(1 for b in draft["_blockDrafts"] if b["skipped"])}개) → {args.out}')


if __name__ == '__main__':
    main()
